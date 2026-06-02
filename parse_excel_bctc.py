import openpyxl
import pandas as pd
from datetime import datetime
import logging

# Cau hinh logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)


class VietStockExcelParser:
    """
    Parser de lay du lieu tai chinh tu file Excel cua VietStock
    Lay 4 chi so: Asset, Doanh thu, LNST, VCSH
    """
    
    def __init__(self, excel_file):
        """Khoi tao parser"""
        self.excel_file = excel_file
        self.wb = openpyxl.load_workbook(excel_file)
        self.data = []
        logger.info(f"✓ Tai file Excel: {excel_file}")
    
    def get_sheet_names(self):
        """Lay danh sach cac sheet trong file"""
        return self.wb.sheetnames
    
    def parse_bctc_tom_tat(self):
        """
        Parse sheet 'BCTC TOM TAT' (Bao cao tai chinh tom tat)
        Sheet nay chua: Asset, Doanh thu, LNST, VCSH
        """
        try:
            sheet = self.wb['BCTC TÓM TẮT']
            logger.info("✓ Dang xu ly sheet: BCTC TOM TAT")
            
            # Hang header co thua nien (nam): 2017, 2018, 2019, 2020, 2021, 2022, 2023, 2024, 2025
            # Vị trí: Hang 6 (index 6), Tu cot D (index 3) tro di
            
            # Tim cac cot tuong ung voi nam 2020-2025
            header_row = 6  # Hang chua nam
            year_columns = {}  # {2020: col_D, 2021: col_E, ...}
            
            for col_idx, cell in enumerate(sheet[header_row], 1):
                if cell.value and isinstance(cell.value, (int, str)):
                    try:
                        year = int(str(cell.value).strip())
                        if 2020 <= year <= 2025:
                            year_columns[year] = col_idx
                    except ValueError:
                        pass
            
            logger.info(f"Tim thay cac nam: {sorted(year_columns.keys())}\")\n\n            # Dong chua cac chi so can lay (Tim bang tim kiem ten)
            # Truong hop chung:
            # - Tong tai san / Total Assets / Tai san (hang 12)
            # - Doanh thu / Revenue (hang tuong ung)
            # - Loi nhuan rong / Net Income / LNST (hang tuong ung)
            # - Von chu so huu / Equity / VCSH (hang tuong ung)
            
            for row_idx, row in enumerate(sheet.iter_rows(min_row=7), 7):\n                col_a = sheet[f'A{row_idx}'].value\n                \n                # Detect: Tai san (Total Assets)\n                if col_a and 'tai san' in str(col_a).lower():\n                    self._extract_year_data(sheet, row_idx, year_columns, 'asset')\n                \n                # Detect: Doanh thu (Revenue)\n                elif col_a and 'doanh thu' in str(col_a).lower():\n                    self._extract_year_data(sheet, row_idx, year_columns, 'revenue')\n                \n                # Detect: LNST (Net Income)\n                elif col_a and ('lnst' in str(col_a).lower() or 'loi nhuan rong' in str(col_a).lower()):\n                    self._extract_year_data(sheet, row_idx, year_columns, 'net_income')\n                \n                # Detect: VCSH (Equity)\n                elif col_a and ('vcsh' in str(col_a).lower() or 'von chu so huu' in str(col_a).lower()):\n                    self._extract_year_data(sheet, row_idx, year_columns, 'equity')\n        \n        except KeyError as e:\n            logger.error(f\"Sheet khong tim thay: {e}\")\n        except Exception as e:\n            logger.error(f\"Loi khi parse BCTC TOM TAT: {e}\")\n    \n    def _extract_year_data(self, sheet, row_idx, year_columns, metric_name):\n        \"\"\"\n        Trich xuat du lieu cho cac nam tu mot hang\n        \"\"\"\n        for year, col_idx in year_columns.items():\n            cell_value = sheet.cell(row=row_idx, column=col_idx).value\n            \n            try:\n                if cell_value:\n                    value = float(cell_value)\n                    self.data.append({\n                        'year': year,\n                        'metric': metric_name,\n                        'value': value\n                    })\n                    logger.info(f\"  {metric_name} ({year}): {value}\")\n            except (ValueError, TypeError):\n                pass\n    \n    def parse_kqkd(self):\n        \"\"\"\n        Parse sheet 'Ket qua kinh doanh' (Operational Results)\n        Chua them chi tiet ve Doanh thu, Chi phi, LNST\n        \"\"\"\n        try:\n            sheet = self.wb['KQKD']\n            logger.info(\"✓ Dang xu ly sheet: KQKD\")\n            \n            # Tuong tu nhu parse_bctc_tom_tat\n            # Tuy nhien, file co the khong co sheet nay\n        except KeyError:\n            logger.warning(\"Sheet 'KQKD' khong tim thay\")\n        except Exception as e:\n            logger.error(f\"Loi khi parse KQKD: {e}\")\n    \n    def get_dataframe(self):\n        \"\"\"\n        Tra ve DataFrame cua du lieu da trich xuat\n        \"\"\"\n        return pd.DataFrame(self.data)\n    \n    def save_to_csv(self, output_file='financial_data.csv'):\n        \"\"\"\n        Luu du lieu ra file CSV\n        \"\"\"\n        df = self.get_dataframe()\n        df.to_csv(output_file, index=False, encoding='utf-8-sig')\n        logger.info(f\"✓ Du lieu da luu: {output_file}\")\n        return output_file\n    \n    def display_data(self):\n        \"\"\"\n        Hien thi du lieu da trich xuat\n        \"\"\"\n        df = self.get_dataframe()\n        print(\"\\n\" + \"=\"*80)\n        print(\"DU LIEU TAI CHINH (2020-2025)\")\n        print(\"=\"*80)\n        print(df.to_string(index=False))\n        print(\"\\nTong records:\", len(df))\n        print(\"=\"*80 + \"\\n\")\n\n\nif __name__ == '__main__':\n    # Duong dan file Excel\n    excel_file = 'BCTC_2026.xlsx'  # Thay bang ten file cua ban\n    \n    try:\n        parser = VietStockExcelParser(excel_file)\n        \n        # Lay danh sach cac sheet\n        print(\"\\nCac sheet trong file:\")\n        for sheet_name in parser.get_sheet_names():\n            print(f\"  - {sheet_name}\")\n        \n        # Parse du lieu\n        parser.parse_bctc_tom_tat()\n        \n        # Hien thi va luu du lieu\n        parser.display_data()\n        parser.save_to_csv('financial_data_extracted.csv')\n        \n    except FileNotFoundError:\n        logger.error(f\"Khong tim thay file: {excel_file}\")\n    except Exception as e:\n        logger.error(f\"Loi: {e}\")\n